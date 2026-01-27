import openpyxl
from openpyxl import load_workbook

def get_row_count(file, sheet_name):
    wb = load_workbook(file)
    sheet = wb[sheet_name]
    row_count = sheet.max_row
    wb.close()
    return row_count

def get_column_count(file, sheet_name):
    wb = load_workbook(file)
    sheet = wb[sheet_name]
    col_count = sheet.max_column
    wb.close()
    return col_count

def read_cell(file, sheet_name, row, column):
    wb = load_workbook(file)
    sheet = wb[sheet_name]
    value = sheet.cell(row=row, column=column).value
    wb.close()
    return value

def write_cell(file, sheet_name, row, column, value):
    wb = load_workbook(file)
    sheet = wb[sheet_name]
    sheet.cell(row=row, column=column, value=value)
    wb.save(file)
    wb.close()

def get_data_as_list(file, sheet_name, skip_header=True):
    """
    Reads all rows in the sheet to a list of tuples.
    If skip_header=True, skips the first row (column names).
    Useful for data-driven tests.
    """
    wb = load_workbook(file)
    sheet = wb[sheet_name]
    data = []
    start_row = 2 if skip_header else 1  # Excel is 1-indexed
    for i in range(start_row, sheet.max_row + 1):
        row_data = []
        for j in range(1, sheet.max_column + 1):
            val = sheet.cell(row=i, column=j).value
            row_data.append(val)
        data.append(tuple(row_data))
    wb.close()
    return data

def get_test_data_dict(file, sheet_name):
    """
    Returns test data as a list of dictionaries with column headers as keys.
    More readable for test data management.
    """
    wb = load_workbook(file)
    sheet = wb[sheet_name]
    
    # Get headers from first row
    headers = []
    for j in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=j).value
        headers.append(header)
    
    # Get data rows
    data = []
    for i in range(2, sheet.max_row + 1):
        row_dict = {}
        for j in range(1, sheet.max_column + 1):
            val = sheet.cell(row=i, column=j).value
            row_dict[headers[j-1]] = val
        data.append(row_dict)
    
    wb.close()
    return data

def update_test_result(file, sheet_name, username, result, result_column=4):
    """
    Helper function to update test results in Excel file.
    Finds the row with matching username and updates the result column.
    """
    wb = load_workbook(file)
    sheet = wb[sheet_name]
    
    for i in range(2, sheet.max_row + 1):  # Start from row 2 (skip header)
        excel_username = sheet.cell(row=i, column=1).value
        if excel_username == username:
            sheet.cell(row=i, column=result_column, value=result)
            wb.save(file)
            wb.close()
            return True
    
    wb.close()
    return False
